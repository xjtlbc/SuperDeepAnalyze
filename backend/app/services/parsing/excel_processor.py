"""Enriched Excel processor using pandas + openpyxl.

Bypasses Docling entirely for Excel files. Produces:
  - L2 markdown with multi-sheet preservation, formula annotations,
    comment annotations, hyperlink annotations, number formatting,
    and merged-cell handling.
  - analysis JSON with column classification, distributions, and findings.
"""

import logging
import re
import statistics
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path

import openpyxl
import pandas as pd

logger = logging.getLogger("app.parsing")


@dataclass
class BannerInfo:
    """Detected banner row in a sheet."""
    row_index: int          # 0-based row index in df_raw
    text: str               # banner text content


@dataclass
class SheetAnalysis:
    name: str
    rows: int
    columns: int
    col_info: list[dict] = field(default_factory=list)
    distributions: list[dict] = field(default_factory=list)
    findings: list[dict] = field(default_factory=list)
    sample_rows: list[list] = field(default_factory=list)
    banners: list[BannerInfo] = field(default_factory=list)


@dataclass
class ExcelProcessResult:
    l2_markdown: str
    analysis: dict


# Number format pattern helpers
_DATE_FORMAT_RE = re.compile(r"(y{2,4}|m{1,5}|d{1,4})", re.IGNORECASE)
_PERCENT_RE = re.compile(r"0*%")
_CURRENCY_RE = re.compile(r"[¥$€£]")


def _classify_column(values: list[str]) -> str:
    """Classify a column's data type from its non-empty string values."""
    if not values:
        return "empty"

    num_count = 0
    date_count = 0
    bool_count = 0
    total = len(values)

    for v in values:
        v = v.strip()
        if not v:
            continue
        # Boolean check
        if v.lower() in ("true", "false", "是", "否", "对", "错", "1", "0"):
            bool_count += 1
            continue
        # Number check
        try:
            float(v.replace(",", ""))
            num_count += 1
            continue
        except ValueError:
            pass
        # Date check
        for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y年%m月%d日"):
            try:
                datetime.strptime(v, fmt)
                date_count += 1
                break
            except ValueError:
                continue

    non_empty = total - sum(1 for v in values if not v.strip())
    if non_empty == 0:
        return "empty"
    if bool_count / non_empty > 0.8:
        return "boolean"
    if num_count / non_empty > 0.8:
        return "float" if any("." in v for v in values if v.strip()) else "integer"
    if date_count / non_empty > 0.5:
        return "date"
    return "string"


def _compute_distribution(values: list[str], col_name: str, dtype: str) -> dict:
    """Compute distribution stats for a column."""
    non_empty = [v.strip() for v in values if v and v.strip()]
    null_count = len(values) - len(non_empty)

    result = {"column": col_name, "type": dtype, "nullCount": null_count, "uniqueCount": len(set(non_empty))}

    if dtype in ("integer", "float"):
        nums = []
        for v in non_empty:
            try:
                nums.append(float(v.replace(",", "")))
            except ValueError:
                pass
        if nums:
            result["stats"] = {
                "mean": round(sum(nums) / len(nums), 2),
                "median": round(statistics.median(nums), 2),
                "std": round(statistics.stdev(nums), 2) if len(nums) > 1 else 0,
                "min": min(nums),
                "max": max(nums),
            }
    elif dtype == "string" and non_empty:
        from collections import Counter
        counts = Counter(non_empty)
        top = counts.most_common(15)
        total = len(non_empty)
        result["stats"] = {
            "topValues": [{"value": v, "count": c, "percentage": round(c / total * 100, 1)} for v, c in top]
        }
    elif dtype == "date":
        dates = []
        for v in non_empty:
            for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"):
                try:
                    dates.append(datetime.strptime(v, fmt))
                    break
                except ValueError:
                    continue
        if dates:
            result["stats"] = {"earliest": min(dates).strftime("%Y-%m-%d"), "latest": max(dates).strftime("%Y-%m-%d")}

    return result


def _detect_findings(col_info: list[dict], total_rows: int) -> list[dict]:
    """Auto-generate data quality findings."""
    findings = []
    for ci in col_info:
        # Constant column
        if ci.get("uniqueCount") == 1 and total_rows > 5:
            findings.append({
                "type": "constant_column",
                "column": ci["name"],
                "detail": f"所有 {total_rows} 行值相同，无区分度",
            })
        # High null fraction
        if ci.get("nullCount", 0) > total_rows * 0.5:
            pct = round(ci["nullCount"] / total_rows * 100, 1)
            findings.append({
                "type": "high_null_fraction",
                "column": ci["name"],
                "detail": f"空值率 {pct}% ({ci['nullCount']}/{total_rows})",
            })
        # High cardinality (near-primary-key)
        if ci.get("uniqueCount", 0) > total_rows * 0.9 and total_rows > 10:
            findings.append({
                "type": "high_cardinality",
                "column": ci["name"],
                "detail": f"唯一值 {ci['uniqueCount']} 接近总行数 {total_rows}（可能是主键）",
            })
    return findings


def _detect_banner_rows(ws, df_raw: pd.DataFrame) -> list[BannerInfo]:
    """Detect banner rows: merged cells spanning the table width where 3+
    consecutive cells have identical text covering >= 40% of non-empty cells.

    This pattern is common in Chinese government/enterprise Excel files where
    a merged row at the top serves as a section title (e.g. "中央党群机关").

    Algorithm (from hc/DeepAnalyze _fix_merged_banner_rows):
    1. For each row, check if it contains a wide merged region (>= 40% of columns).
    2. Within that row, find groups of consecutive cells with identical non-empty text.
    3. If a group spans >= 3 cells and covers >= 40% of non-empty cells in the row,
       treat it as a banner row.
    """
    if df_raw.empty:
        return []

    rows, cols = df_raw.shape
    banners: list[BannerInfo] = []

    # Build a map: (row_num, col_num) -> merged cell range for quick lookup
    merged_at: dict[tuple[int, int], openpyxl.worksheet.cell_range.CellRange] = {}
    for mr in ws.merged_cells.ranges:
        for row in range(mr.min_row, mr.max_row + 1):
            for col in range(mr.min_col, mr.max_col + 1):
                merged_at[(row, col)] = mr

    # Check at most the first 5 rows for banners (header is usually near top)
    max_check = min(rows, 5)

    for df_row_idx in range(max_check):
        excel_row = df_row_idx + 1  # 1-based

        # Collect cell values for this row
        cell_values: list[str] = []
        for col_idx in range(cols):
            val = df_raw.iloc[df_row_idx, col_idx]
            cell_values.append(str(val).strip() if pd.notna(val) else "")

        non_empty_count = sum(1 for v in cell_values if v)
        if non_empty_count == 0:
            continue

        # Check for wide merged region in this row
        has_wide_merge = False
        merge_col_span = 0
        for col_idx in range(cols):
            mr = merged_at.get((excel_row, col_idx + 1))
            if mr and mr.min_row == excel_row:
                span = mr.max_col - mr.min_col + 1
                if span >= cols * 0.4:  # merged region spans >= 40% of table width
                    has_wide_merge = True
                    merge_col_span = max(merge_col_span, span)

        if not has_wide_merge:
            continue

        # Find groups of consecutive cells with identical non-empty text
        groups: list[tuple[str, int, int]] = []  # (text, start_idx, end_idx)
        i = 0
        while i < cols:
            v = cell_values[i]
            if not v:
                i += 1
                continue
            # Find run of identical text
            j = i + 1
            while j < cols and cell_values[j] == v:
                j += 1
            if j - i >= 1:
                groups.append((v, i, j))
            i = j

        # Check if any group qualifies as a banner
        for text, start, end in groups:
            span = end - start
            if span >= 3 and span >= non_empty_count * 0.4:
                banners.append(BannerInfo(row_index=df_row_idx, text=text))
                break  # Only one banner per row

    return banners


def _format_number(value, fmt_str: str) -> str:
    """Apply Excel number format to a value."""
    if not fmt_str or fmt_str == "General":
        return str(value) if value is not None else ""

    # Date format
    if _DATE_FORMAT_RE.search(fmt_str):
        if isinstance(value, (int, float)):
            # Excel date serial number
            try:
                dt = datetime.fromordinal(int(value) + 693594)
                # Excel serial offset for 1900 date system
                if value > 59:
                    pass  # Lotus bug already handled by fromordinal offset
                return dt.strftime("%Y-%m-%d")
            except (ValueError, OverflowError):
                pass
        elif isinstance(value, datetime):
            return value.strftime("%Y-%m-%d")

    # Percentage
    if _PERCENT_RE.search(fmt_str):
        try:
            return f"{float(value) * 100:.1f}%"
        except (ValueError, TypeError):
            pass

    # Currency
    if _CURRENCY_RE.search(fmt_str):
        try:
            return f"¥{float(value):,.2f}" if "¥" in fmt_str else f"${float(value):,.2f}"
        except (ValueError, TypeError):
            pass

    return str(value) if value is not None else ""


def _analyze_sheet(ws, sheet_name: str, df_raw: pd.DataFrame) -> SheetAnalysis:
    """Analyze a single worksheet: column classification, distributions, findings."""
    rows, cols = df_raw.shape
    analysis = SheetAnalysis(name=sheet_name, rows=rows, columns=cols)

    if rows == 0 or cols == 0:
        return analysis

    # Detect banner rows before proceeding
    banners = _detect_banner_rows(ws, df_raw)
    analysis.banners = banners

    # Determine the effective header row, skipping banners
    banner_indices = {b.row_index for b in banners}
    header_row_idx = 0
    while header_row_idx in banner_indices and header_row_idx < rows:
        header_row_idx += 1

    if header_row_idx >= rows:
        return analysis

    # Use the effective header row
    headers = [str(df_raw.iloc[header_row_idx, c]) if pd.notna(df_raw.iloc[header_row_idx, c]) else f"Col_{c}" for c in range(cols)]
    data_values = {c: [str(df_raw.iloc[r, c]) if pd.notna(df_raw.iloc[r, c]) else "" for r in range(rows) if r > header_row_idx and r not in banner_indices]
                   for c in range(cols)}

    # Classify columns
    col_info = []
    for c in range(cols):
        values = data_values[c]
        dtype = _classify_column(values)
        unique = len(set(v for v in values if v.strip()))
        sample = [v for v in values if v.strip()][:5]
        col_info.append({
            "name": headers[c],
            "dataType": dtype,
            "nullCount": sum(1 for v in values if not v.strip()),
            "uniqueCount": unique,
            "sampleValues": sample,
        })
    analysis.col_info = col_info

    # Distributions
    data_row_count = rows - 1
    for ci in col_info:
        if ci["dataType"] == "empty":
            continue
        values = data_values[headers.index(ci["name"])]
        analysis.distributions.append(_compute_distribution(values, ci["name"], ci["dataType"]))

    # Findings
    analysis.findings = _detect_findings(col_info, data_row_count)

    # Sample rows (first 3 data rows, skipping banners)
    sample_data_rows = [r for r in range(rows) if r > header_row_idx and r not in banner_indices]
    sample_rows = sample_data_rows[:3]
    analysis.sample_rows = [
        [str(df_raw.iloc[r, c]) if pd.notna(df_raw.iloc[r, c]) else "" for c in range(cols)]
        for r in sample_rows
    ]

    return analysis


def _generate_sheet_markdown(ws, df: pd.DataFrame, header_row_idx: int,
                             banners: list[BannerInfo] | None = None) -> str:
    """Generate enriched Markdown table for a worksheet.

    Preserves: formulas [fx:], comments [note:], hyperlinks [link:],
    number formatting, merged-cell handling.
    Banner rows are emitted as ``## Banner: {text}`` sections before the table
    and excluded from the data table itself.
    """
    if df.shape[0] == 0 or df.shape[1] == 0:
        return ""

    banner_indices = {b.row_index for b in (banners or [])}

    # Build set of merged cell positions (non-top-left)
    merged_shadow = set()
    for mr in ws.merged_cells.ranges:
        for row in range(mr.min_row, mr.max_row + 1):
            for col in range(mr.min_col, mr.max_col + 1):
                if row != mr.min_row or col != mr.max_col:
                    merged_shadow.add((row, col))

    # Header row index (0-based in df, but Excel rows are 1-based)
    # Excel row of first data row = (header_row_idx + 1) [skip header] + 1 [1-based] = header_row_idx + 2
    # But df rows include banner rows before the header, so we must count skipped banner rows
    header_offset = header_row_idx + 2  # Excel row number of first data row

    rows_md = []
    cols_count = df.shape[1]

    for df_row_idx in range(df.shape[0]):
        # Skip banner rows — they are emitted separately above the table
        if df_row_idx in banner_indices:
            continue

        # Adjust for skipped banner rows between df start and this row
        skipped_before = sum(1 for bi in banner_indices if bi < df_row_idx and bi != header_row_idx)
        excel_row = df_row_idx + header_offset - skipped_before
        cells = []

        for col_idx in range(cols_count):
            excel_col = col_idx + 1
            cell_key = (excel_row, excel_col)

            # Skip merged shadow cells
            if cell_key in merged_shadow:
                cells.append("")
                continue

            cell = ws.cell(row=excel_row, column=excel_col)
            raw_val = df.iloc[df_row_idx, col_idx] if pd.notna(df.iloc[df_row_idx, col_idx]) else ""

            # Apply number formatting
            display_val = raw_val
            if cell.number_format and cell.number_format != "General" and raw_val != "":
                display_val = _format_number(raw_val, cell.number_format)
            else:
                display_val = str(raw_val) if raw_val != "" else ""

            # Collect annotations
            annotations = []
            if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                annotations.append(f"[fx:{cell.value}]")
            elif raw_val != "" and hasattr(cell, "value") and isinstance(cell.value, str) and cell.value.startswith("="):
                annotations.append(f"[fx:{cell.value}]")

            if cell.comment:
                comment_text = cell.comment.text.replace("\n", " ").strip()[:50]
                annotations.append(f"[note:{comment_text}]")

            if cell.hyperlink and cell.hyperlink.target:
                annotations.append(f"[link:{cell.hyperlink.target}]")

            cell_str = display_val + (" " + " ".join(annotations) if annotations else "")
            cells.append(cell_str)

        rows_md.append(cells)

    if not rows_md:
        return ""

    # Build banner sections (before the table)
    banner_lines = []
    for b in sorted(banners or [], key=lambda b: b.row_index):
        banner_lines.append(f"## Banner: {b.text}")
        banner_lines.append("")

    # Build markdown table
    lines = list(banner_lines)
    # Header
    lines.append("| " + " | ".join(str(c) for c in rows_md[0]) + " |")
    lines.append("| " + " | ".join(["---"] * cols_count) + " |")
    # Data rows
    for row in rows_md[1:]:
        lines.append("| " + " | ".join(str(c) for c in row) + " |")

    return "\n".join(lines)


def process_excel(file_path: str | Path) -> ExcelProcessResult:
    """Process an Excel file producing L2 markdown and analysis JSON.

    Args:
        file_path: Path to the Excel file (.xlsx, .xls, .xlsm)

    Returns:
        ExcelProcessResult with l2_markdown and analysis dict
    """
    path = Path(file_path)
    logger.info("Processing Excel file: %s", path.name)

    # Load with openpyxl (preserves formulas, comments, hyperlinks)
    wb = openpyxl.load_workbook(str(path), data_only=False)

    # Load with pandas for raw data
    sheet_names = wb.sheetnames

    all_sheet_markdown = []
    analysis_sheets = []

    for sheet_name in sheet_names:
        ws = wb[sheet_name]

        # Read raw data with pandas
        try:
            df_raw = pd.read_excel(path, sheet_name=sheet_name, header=None, dtype=str)
        except Exception as e:
            logger.warning("Failed to read sheet '%s' with pandas: %s", sheet_name, e)
            continue

        if df_raw.empty:
            continue

        # Detect banner rows early (needed for both markdown and analysis)
        banners = _detect_banner_rows(ws, df_raw)
        banner_indices = {b.row_index for b in banners}

        # Detect header row: skip banner rows
        header_row_idx = 0
        while header_row_idx in banner_indices and header_row_idx < df_raw.shape[0]:
            header_row_idx += 1

        # Generate enriched markdown (with banners if detected)
        md = _generate_sheet_markdown(ws, df_raw, header_row_idx, banners=banners)
        if md:
            all_sheet_markdown.append(f"# Sheet: {sheet_name}\n\n{md}")

        # Analyze sheet (reuses banners already detected)
        try:
            analysis = _analyze_sheet(ws, sheet_name, df_raw)
            banners_info = [{"rowIndex": b.row_index, "text": b.text} for b in analysis.banners]

            analysis_sheets.append({
                "name": analysis.name,
                "dimensions": {"rows": analysis.rows, "columns": analysis.columns},
                "columns": analysis.col_info,
                "distributions": analysis.distributions,
                "findings": analysis.findings,
                "banners": banners_info,
            })
        except Exception as e:
            logger.warning("Failed to analyze sheet '%s': %s", sheet_name, e)

    wb.close()

    l2_markdown = "\n\n".join(all_sheet_markdown) if all_sheet_markdown else ""
    analysis = {
        "fileName": path.name,
        "sheets": analysis_sheets,
    }

    logger.info("Excel processed: %d sheets, %d chars markdown", len(analysis_sheets), len(l2_markdown))
    return ExcelProcessResult(l2_markdown=l2_markdown, analysis=analysis)
